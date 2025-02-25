import json
import os
import pandas as pd
import numpy as np
import configparser
import openpyxl
import datetime
import math
import warnings
import paramiko

def loadData(csvtoggle, csvpath, useftp, sftp_client, ftppath):
    df = pd.DataFrame()
    root_dirnames = []
    if useftp == "true":
        with open("../data/usercache.json", "wb") as file:
            sftp_client.getfo("usercache.json", file)
        names = pd.DataFrame(json.load(open("../data/usercache.json", "r")))
        sftp_client.chdir("../")
        
        root_dirnames = sftp_client.listdir(ftppath)
        sftp_client.chdir(ftppath)
        
        for dirname in root_dirnames:
            if dirname[-1] == ".":
                continue
            sftp_client.chdir(dirname)
            filenames = sftp_client.listdir()
            
            for filename in filenames:
                if filename == "." or filename == "..":
                    continue
                print("Now processing", filename)
                
                local_file = f"temp_{filename}"
                sftp_client.get(filename, local_file)
                
                with open(local_file, "r") as file:
                    data = json.load(file)['extraData']['cobbledex_discovery']['registers']
                
                os.remove(local_file)
                
                temp_df = pd.json_normalize(data, meta_prefix=True)
                temp_name = names.loc[names['uuid'] == filename[:-5]]['name']
                temp_df = temp_df.transpose().iloc[:]
                if temp_name.empty:
                    print("No username found for UUID", filename[:-5], " in usercache.json, using UUID for this player instead.")
                    temp_name = filename[:-5]
                    temp_df = temp_df.rename({0: temp_name}, axis=1)
                else:
                    temp_df = temp_df.rename({0: temp_name.iloc[0]}, axis=1)
                
                if not temp_df.empty:
                    temp_df.index = temp_df.index.str.split('.', expand=True)
                    if df.empty:
                        df = temp_df
                    else:
                        df = df.join(temp_df, how="outer")
                else:
                    df[temp_name] = np.nan
                
            sftp_client.chdir("../")
    
    df = df.fillna(0)
    if csvtoggle == "true":
        df.to_csv(csvpath)
    return df

def most_pokemons_leaderboard(df, config, type):
    file_path = "output.xlsx"
    wb = openpyxl.load_workbook(file_path)
    sheet_name = {"standard": "Global", "shiny": "Shiny", "legendary": "Legendary"}[type]
    ws = wb[sheet_name]
    
    i = 0
    ExcelRows = int(config['ExcelRows'])
    ExcelCols = int(config['ExcelColumns'])
    for index, row in df[0:ExcelRows*ExcelCols].iterrows():
        ws.cell(row=(i%ExcelRows)+3, column=2+math.floor(i/ExcelRows)*3, value=str(i+1)+".")
        ws.cell(row=(i%ExcelRows)+3, column=3+math.floor(i/ExcelRows)*3, value=index)
        ws.cell(row=(i%ExcelRows)+3, column=4+math.floor(i/ExcelRows)*3, value=row[0])
        i += 1
    now = datetime.datetime.now()
    ws.cell(row=ExcelRows+3, column=2, value=now.strftime(config['LastUpdated']))
    ws.cell(row=ExcelRows+4, column=2, value=config['Subtitle'])
    wb.save(file_path)

# Read config
config = configparser.ConfigParser()
config.read('cobblemon_config.ini')


sftp_client = None
if config['FTP']['UseFTP'] == "true":
    transport = paramiko.Transport((config['FTP']['Host'], int(config['FTP']['Port'])))
    transport.connect(username=open("../username.txt", "r").read().strip(), password=open("../password.txt", "r").read().strip())
    sftp_client = paramiko.SFTPClient.from_transport(transport)

df = loadData(config['GLOBALMATRIX']['CreateCSV'], config['GLOBALMATRIX']['CSVPath'], config['FTP']['UseFTP'], sftp_client, config['FTP']['Path'])

if config['FTP']['UseFTP'] == "true":
    sftp_client.close()
    transport.close()

count_df = df.drop(['caughtTimestamp', 'discoveredTimestamp', 'isShiny'], level=2)
pokemons_db = pd.read_csv('Pokemon.csv')
legendary_list = pokemons_db.loc[pokemons_db['Legendary'] == True]

if config['LEADERBOARD']['Enable'] == "true":
    player_sum = pd.DataFrame((count_df == "CAUGHT").sum().sort_values())
    player_sum['index'] = range(len(player_sum), 0, -1)
    player_sum = player_sum.iloc[::-1]
    ignore_names = [name.strip() for name in config['LEADERBOARD']['IgnoreNames'].split(",") if name.strip()]
    player_sum.drop(ignore_names, inplace=True, errors='ignore')
    most_pokemons_leaderboard(player_sum, config['LEADERBOARD'], "standard")

if config['SHINYLEADERBOARD']['Enable'] == "true":
    player_sum = pd.DataFrame(((df == "True") | (df == True)).sum().sort_values())
    player_sum['index'] = range(len(player_sum), 0, -1)
    player_sum = player_sum.iloc[::-1]
    ignore_names = [name.strip() for name in config['SHINYLEADERBOARD']['IgnoreNames'].split(",") if name.strip()]
    player_sum.drop(ignore_names, inplace=True, errors='ignore')
    most_pokemons_leaderboard(player_sum, config['SHINYLEADERBOARD'], "shiny")

if config['LEGLEADERBOARD']['Enable'] == "true":
    legs = legendary_list['Cobblemon'].tolist()
    leg_count_df = count_df.loc[count_df.index.get_level_values(0).isin(legs)]
    with warnings.catch_warnings():
        warnings.simplefilter(action='ignore', category=FutureWarning)
        leg_count_df = leg_count_df.groupby(level=0).agg(lambda x: "CAUGHT" if "CAUGHT" in x.values else 0)
    player_sum = pd.DataFrame((leg_count_df == "CAUGHT").sum().sort_values())
    player_sum['index'] = range(len(player_sum), 0, -1)
    player_sum = player_sum.iloc[::-1]
    ignore_names = [name.strip() for name in config['LEGLEADERBOARD']['IgnoreNames'].split(",") if name.strip()]
    player_sum.drop(ignore_names, inplace=True, errors='ignore')
    most_pokemons_leaderboard(player_sum, config['LEGLEADERBOARD'], "legendary")

print("Done!")
